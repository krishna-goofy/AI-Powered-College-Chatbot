from flask import Flask, render_template, request, jsonify
from scraper.scraper import scrape_website, download_pdf
from database.db import create_database, insert_web_data, insert_pdf_data, get_processed_urls, add_processed_url, get_processed_pdfs, add_processed_pdf
import sqlite3
import google.generativeai as genai
import os
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl
from openpyxl import Workbook


app = Flask(__name__)

excel_file_path = "contact_data.xlsx"

def initialize_excel_file():
    """Create an Excel file with headers if it doesn't exist."""
    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contact Form Data"
        ws.append(["Name", "Phone", "Message"])  # Add headers
        wb.save(excel_file_path)

initialize_excel_file()  # Call this function when the app starts

# Create database and trigger initial scraping if database is empty
def check_and_scrape_initial_data():
    create_database()

    # Check if web_data or pdf_data tables are empty
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM web_data")
    web_data_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM pdf_data")
    pdf_data_count = cursor.fetchone()[0]
    conn.close()

    # Trigger scraping if database is empty
    if web_data_count == 0 and pdf_data_count == 0:
        print("Database is empty. Starting initial scraping...")
        url = "https://www.msijanakpuri.com/"  # Your college website
        scrape_and_process_data(url)
        print("Initial scraping completed!")

# Gemini API configuration
genai.configure(api_key="AIzaSyCDo597iianSU8muPOlF_fKomZaFJm1vhk")  # Replace with your Gemini API key

# Gemini API function using google.generativeai
def ask_gemini(question, db_content):
    model = genai.GenerativeModel("gemini-1.5-flash")
    prompt = (f"You are a friendly assistant. Provide a short, engaging, and conversational response to the question below. "
              f"Use the context if needed, but keep the answer brief.\n\n"
              f"Context:\n{db_content}\n\n"
              f"Question: {question}")
    response = model.generate_content(prompt)
    cleaned_text = response.text.replace('*', '')

    return {"answer": cleaned_text}


# Route to render the front end
@app.route('/')
def index():
    return render_template('index.html')

# Route to return predefined questions as JSON
@app.route('/get_predefined_questions', methods=['GET'])
def get_predefined_questions():
    questions = [
        "Contact",
        "Tell me About the College",
        "What is the admission process?",
        "What courses do you offer?",
        "Where is the campus located?",
        "How to apply for scholarships?",
        "What is the fee structure?"
        "What programs/majors are offered?",
        "What are admission requirements?",
        "Can I get information on fees?",
        "What scholarships are available?",
        "How do I check my application status?",
        "What documents are required for admission?",
        "What courses are offered this semester?",
        "Can I get a course catalog/syllabus?",
        "Who is my academic advisor?",
        "How do I register for classes?",
        "What are the academic policies?",
        "Can I get information on research opportunities?",
        "How do I access online course materials?",
        "What student organizations/clubs are available?",
        "How do I access mental health resources?",
        "What dining options are on campus?",
        "Can I get information on campus events?",
        "How do I report a concern/incident?",
        "What support services are available for students with disabilities?",
        "How do I access career counseling?",
        "How do I apply for financial aid?",
        "What payment options are available?",
        "Can I get information on student loans?",
        "How do I access my billing statement?",
        "What scholarships are available?",
        "How do I set up a payment plan?",
        "Can I get a refund/withdrawal information?",
        "Where is the library/learning center?",
        "How do I access IT support?",
        "What recreational facilities are available?",
        "Can I get information on campus safety/security?",
        "How do I report a maintenance issue?",
        "What health services are available?",
        "Can I get information on parking/transportation?",
        "What are the college's mission/values?",
        "Can I get contact information for faculty/staff?",
        "How do I access the college's website/student portal?",
        "What social media platforms does the college use?",
        "Can I get information on campus tours?",
        "How do I provide feedback/suggestions?",
        "Are there any upcoming events?",
        "I'm interested in a program, can you tell me more?",
        "I'm struggling with an academic/personal issue, can you offer support?",
        "Can you recommend resources for a specific topic?",
        "How do I get involved in campus activities?",
        "Can you walk me through the [process/procedure]?"
    ]
    return jsonify(questions)

# Route to trigger scraping
@app.route('/scrape', methods=['POST'])
def scrape():
    url = request.json.get('url')
    scrape_and_process_data(url)
    return jsonify({"message": "Scraping completed!"})

# Function to scrape website and process data
def scrape_and_process_data(url):
    processed_urls = get_processed_urls()  # Get already processed URLs
    text, pdf_links = scrape_website(url)

    # Only insert new data into the database
    if url not in processed_urls:
        insert_web_data(url, text)
        add_processed_url(url)  # Add the URL to the processed list

    # Handle PDFs
    processed_pdfs = get_processed_pdfs()  # Get already processed PDFs
    for pdf_url in pdf_links:
        if pdf_url not in processed_pdfs:
            pdf_path, pdf_text = download_pdf(pdf_url)
            if pdf_path:
                insert_pdf_data(pdf_path, pdf_text)
                add_processed_pdf(pdf_url)  # Mark the PDF as processed

@app.route('/ask', methods=['POST'])
def ask_question():
    question = request.json.get('question')

    # Fetch all content from the database
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    
    # Fetch web data
    cursor.execute("SELECT text_data FROM web_data")
    web_data = " ".join([row[0] for row in cursor.fetchall() if row[0] is not None])
    
    # Fetch pdf data
    cursor.execute("SELECT text_data FROM pdf_data")
    pdf_data = " ".join([row[0] for row in cursor.fetchall() if row[0] is not None])
    
    db_content = web_data + " " + pdf_data
    conn.close()

    if "contact" == question.lower():
        return jsonify({"answer": "Do you wish to contact us?", "show_buttons": True})

    # Handle other questions using the Gemini API
    response = ask_gemini(question, db_content)

    return jsonify(response)



@app.route('/contact', methods=['POST'])
def contact():
    name = request.json.get('name')
    phone = request.json.get('phone')
    message = request.json.get('message')

    # Save contact details to Excel
    wb = Workbook()
    wb = openpyxl.load_workbook(excel_file_path)  # Load the existing workbook
    ws = wb.active
    ws.append([name, phone, message])  # Append the new data
    wb.save(excel_file_path)  # Save the workbook

    return jsonify({"message": "Contact details submitted successfully!"})


# Scheduler to scrape every hour
def schedule_scraping():
    print("Scheduled scraping started")
    url = "https://www.msijanakpuri.com/"  # Add your college website
    scrape_and_process_data(url)
    print("Scheduled scraping completed")

# Set up background scheduler for periodic scraping
scheduler = BackgroundScheduler()
scheduler.add_job(schedule_scraping, 'interval', hours=1)
scheduler.start()

if __name__ == '__main__':
    try:
        # Check if database is empty and start scraping if necessary
        check_and_scrape_initial_data()

        app.run(host='0.0.0.0', port=5000, debug=True)  # Bind to all interfaces on port 5000
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()
